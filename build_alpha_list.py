#!/usr/bin/env python3
"""Build Revive alpha-customer list from Apollo-verified contacts (July 2026)."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# name, title, company, domain, hq, headcount, email, email_status, what_they_do, why_they_fit, hook, tier
ROWS = [
    # ---------- TIER 1: agent runs on the END USER's OAuth grant. Token dies -> product dies. ----------
    ("Pratik S.", "Founder & CEO", "Tensol", "tensol.ai", "San Francisco", 1, "pratik@tensol.ai", "Verified",
     "AI employees with their own email address, phone number, and tool access",
     "Every 'AI employee' holds a live OAuth grant per tool. One expiry = that employee silently stops working mid-task and nobody notices for a day.",
     "Your AI employee has an inbox and a phone. When its Google grant expires mid-task, does it tell anyone — or just go quiet? Revive catches it, gets the human to re-approve, and resumes the exact step.", "1"),

    ("Zach Zhong", "Co-Founder", "Bubble Lab", "bubblelab.ai", "San Francisco", 3, "zachzhong@bubblelab.ai", "Verified",
     "AI work automation for Slack teams; natural-language workflow builder",
     "Slack workspace token + every downstream tool the workflow touches. A long workflow is exactly the thing that dies halfway.",
     "A natural-language workflow that touches 5 tools has 5 ways to die halfway. We resume from the step that broke instead of re-running the whole thing.", "1"),

    ("Perbhat K.", "Founder", "Glue", "buildwithglue.com", "San Francisco", 2, "perbhat@buildwithglue.com", "Verified",
     "AI brand operations on autopilot (Figma / design / dev tools)",
     "Autopilot means unattended. Unattended is precisely when nobody is around to click 'reconnect'.",
     "'On autopilot' and 'please reconnect your account' can't both be true. Revive is the part that keeps autopilot honest.", "1"),

    ("Paola Martinez", "Co-Founder", "Cofia", "cofia.ai", "New York", 2, "paola@cofia.ai", "Verified",
     "AI automations that observe existing workflows",
     "Observing a customer's workflow means holding read grants across their whole stack — a broad surface of things that expire.",
     "You watch a customer's tools to learn their workflow. What happens to the automations you built when one of those grants dies?", "1"),

    ("Lance Y.", "Founder", "Traverse", "traverse.so", "San Francisco", 3, "lance@traverse.so", "Verified",
     "AI workflow discovery across a company's existing tools",
     "Discovery across N tools = N token lifecycles. Coverage gaps look like 'the product missed something' rather than 'a token died'.",
     "When one connector goes stale, your discovery silently gets less complete — and it reads as a product problem, not an auth problem.", "1"),

    ("Naman A.", "Founder & CEO", "Oximy", "oximy.com", "San Francisco", 4, "naman@oximy.com", "Verified",
     "AI adoption engine for enterprise teams",
     "Enterprise workspace grants, revoked or expired by IT policy on a schedule you don't control.",
     "Enterprise IT rotates and revokes on their calendar, not yours. Revive turns that from an outage into a re-approval link.", "1"),

    ("Arushi G.", "CEO / Co-Founder", "Ressl AI", "ressl.ai", "San Francisco", 5, "arushi@ressl.ai", "Verified",
     "AI employees for trades and home-services companies; event-triggered background work",
     "Your users are plumbers and electricians, not IT admins. They will never notice a 'reconnect your account' banner.",
     "Your users are in a truck, not in a dashboard. When their token dies, they won't see the reconnect banner — but they will see the missed job.", "1"),

    ("Abhishek E.", "Co-Founder", "Ressl AI", "ressl.ai", "San Francisco", 5, "abhishek@ressl.ai", "Verified",
     "AI employees for trades and home-services companies",
     "Second contact at Ressl — CTO-side. Same failure mode, technical framing.",
     "Technical version of the Ressl pitch: refresh-token death mid-run, checkpointed resume, no full re-execution.", "1"),

    ("Shourya J.", "Co-Founder & CEO", "RamAIn", "ramain.ai", "San Francisco", 6, "shourya@ramain.ai", "Verified",
     "AI UI automation for manual ops tasks",
     "UI automation = browser sessions, which die faster and more silently than OAuth tokens. Session-death IS their product surface.",
     "UI automation dies on session expiry more often than anything else in the stack. That's the exact thing we resume — from the step, not the start.", "1"),

    ("Vansh R.", "Co-Founder & CTO", "RamAIn", "ramain.ai", "San Francisco", 6, "vansh@ramain.ai", "Verified",
     "AI UI automation for manual ops tasks",
     "Second contact at RamAIn — the person who actually wrote the session-retry logic.",
     "You've already written session-retry logic. The part nobody writes is the human-in-the-loop re-auth + exact-step resume. That's us.", "1"),

    ("Timothy Chung", "Co-Founder & CEO", "Unifold", "unifold.io", "New York", 3, "tim@unifold.io", "Verified",
     "Workflow-to-API automation for portals (portals with no API)",
     "Portals = login walls + sessions that expire aggressively. Their entire product lives on the wrong side of a session boundary.",
     "You turn portals into APIs. Portals fight back with session expiry and re-login walls. Revive pauses the run, gets a human to sign back in, and resumes.", "1"),

    ("Hau Chu", "Co-Founder", "Unifold", "unifold.io", "New York", 3, "hau@unifold.io", "Verified",
     "Workflow-to-API automation for portals",
     "Second contact at Unifold.",
     "Technical framing of the Unifold pitch: session death on portal automation, checkpoint + resume.", "1"),

    ("Vivek Raja", "Co-Founder & CEO", "Terminal Use", "terminaluse.com", "United Kingdom", 4, "vraja@terminaluse.com", "Verified",
     "Self-improving AI agents / agent platform",
     "Long-horizon agent runs are the single worst case for token expiry — the longer the run, the higher the odds a grant dies inside it.",
     "Self-improving agents run long. The longer the run, the higher the chance a token dies inside it. We make that a pause, not a failure.", "1"),

    ("Stavros Filosidis", "Founder", "Terminal Use", "terminaluse.com", "United States", 4, "sfilosidis@terminaluse.com", "Verified",
     "Self-improving AI agents", "Second contact at Terminal Use.",
     "Same as above, founder-to-founder framing.", "1"),

    ("Filip B.", "Co-Founder", "Terminal Use", "terminaluse.com", "San Francisco", 4, "fbalucha@terminaluse.com", "Verified",
     "Self-improving AI agents", "Third contact at Terminal Use — strongest single fit in the list, worth multiple touches.",
     "Technical framing.", "1"),

    ("Vedant S.", "Co-Founder", "Salus", "usesalus.ai", "Stanford", 2, "vedant@usesalus.ai", "Verified",
     "Policy-aware runtime for AI agents",
     "PARTNER as much as customer. A policy runtime needs a story for 'the agent's credential just died mid-policy-check'.",
     "You own the policy layer for agent runs. We own what happens when the credential under that policy dies mid-run. Complementary — worth 20 min.", "1"),

    ("Alisa Rae", "CEO", "Lucent", "lucenthq.com", "Australia", 5, "alisa@lucenthq.com", "Verified",
     "Observability for AI agents",
     "PARTNER. They can see the run died; they can't resume it. That gap is literally the Revive pitch.",
     "You show people their agent run died. We're the part that brings it back. Observability + recovery is a much better demo than either alone.", "1"),

    ("Ali K.", "Co-Founder & CEO", "Piris Labs", "pirislabs.io", "San Francisco", 6, "alik@pirislabs.io", "Verified",
     "Evals and observability for AI agents",
     "PARTNER. Same as Lucent — they measure failure, we recover from it.",
     "Your evals will flag auth-death as a failed run. It isn't a model failure — it's a recoverable one. Want to compare notes?", "1"),

    # ---------- TIER 2: GTM / CRM agents (Gmail + CRM OAuth, non-technical end users) ----------
    ("Akash P.", "Co-Founder & CEO", "Chasi AI", "chasi.co", "San Francisco", 4, "akash@chasi.co", "Verified",
     "AI concierge for equipment sales and rentals",
     "Customer-facing agent on the customer's inbox/CRM grant. Token death = leads dropped in silence.",
     "An AI concierge that stops answering because a token expired doesn't look like an auth bug to your customer — it looks like your product broke.", "2"),

    ("Devi J.", "Co-Founder", "Cardinal", "trycardinal.ai", "San Francisco", 8, "devi@trycardinal.ai", "Verified",
     "Real-time sales intelligence for GTM teams",
     "Real-time means a stale connector is worse than no connector — you serve confidently wrong data.",
     "'Real-time' + a dead CRM token = confidently stale intelligence. That's worse than an error message.", "2"),

    ("Varun A.", "Founder", "Envariant", "envariant.ai", "San Francisco", 1, "varun@envariant.ai", "Verified",
     "AI for sales research", "Solo founder, deep OAuth surface, will absolutely read a technical cold email.",
     "Solo founder shipping fast — you don't want to spend a week building reauth infra. That's what we are.", "2"),

    ("Victor L.", "CEO & Co-Founder", "Perfectly", "perfectly.so", "San Francisco", 5, "victor@perfectly.so", "Verified",
     "AI deal sourcing for acquisition teams",
     "Deal sourcing runs are long and unattended; a mid-run token death silently truncates the pipeline.",
     "A dead token mid-sourcing-run doesn't error — it just returns fewer deals. Nobody notices until the pipeline looks thin.", "2"),

    ("Varun M.", "Co-Founder & CEO", "Unisson", "unisson.ai", "San Francisco", 2, "varun@unisson.ai", "Verified",
     "AI agents for customer-facing teams",
     "Customer-facing = the failure is visible to YOUR customer's customer. Highest-stakes version of token death.",
     "When your agent's token dies, the person who notices is your customer's customer. That's the worst place to find out.", "2"),

    ("Ketan A.", "Co-Founder", "Carson AI", "usecarson.com", "San Francisco", 4, "ketan@usecarson.com", "Verified",
     "AI-native medspa software", "Practice-management + calendar grants held by non-technical clinic staff.",
     "Your users are front-desk staff at a medspa. They are never going to debug an expired token — but they will churn over a missed booking.", "2"),

    ("Bryan Z.", "Co-Founder", "Reframe", "usereframe.ai", "San Francisco", 1, "bryan@usereframe.ai", "Verified",
     "AI customer text layer for DTC brands",
     "Shopify OAuth + messaging grants. Shopify tokens expire and Shopify's own forums are full of this complaint.",
     "Shopify's own community threads are full of 'my background job died when the offline token expired.' We fixed that.", "2"),

    ("Nolan R.", "CEO", "Scout Out / Foreman", "scoutout.ai", "San Francisco", 4, "nolan@scoutout.ai", "Verified",
     "AI estimating + CRM for contractors",
     "Contractors as end users — the least likely humans on earth to re-authorize a background integration.",
     "Contractors will not click 'reconnect.' Ever. Revive sends the re-approval to whoever will, and resumes the run behind it.", "2"),

    ("Zach Z. (Sparkles)", "CEO", "Sparkles", "sparkles.dev", "San Francisco", 2, "dan@sparkles.dev", "Verified — NAME MISMATCH, check before sending",
     "AI-native recruiting OS for startups",
     "ATS + Gmail + calendar OAuth, all held by a recruiter. Classic silent-death profile.",
     "Recruiting agents live on Gmail + calendar grants. When one dies mid-pipeline, candidates fall through and it looks like your bug.", "2"),

    # ---------- TIER 3: finance / accounting agents (QuickBooks, banking, billing) ----------
    ("Iqbol Temirkhojaev", "Founder & CEO", "Wayco", "wayco.ai", "New York", 2, "iqbol@wayco.ai", "Verified",
     "AI finance teammate for accounting teams",
     "QuickBooks/Xero/banking grants. A dead token mid-close = a broken month-end, which is a fireable offense for their user.",
     "A finance agent that dies mid-close doesn't just fail — it leaves the books half-done. That's the failure your users can't forgive.", "3"),

    ("Thomas Dowling", "Co-Founder & CEO", "FullSeam", "fullseam.com", "United States", 4, "tom@fullseam.com", "Verified",
     "AI agents across accounting, CRM, billing and banking systems",
     "Four token lifecycles per customer, all in one workflow. Highest per-customer breakage odds in the list.",
     "Accounting + CRM + billing + banking = four ways for one workflow to die. We resume from the step, not the start.", "3"),

    ("Mathias Lovring", "Co-Founder & CEO", "Balance", "getbalance.ai", "United Kingdom", 8, "mathias@getbalance.ai", "Verified",
     "Fractional finance team / AI accounting",
     "Same as Wayco — long reconciliation runs on expiring grants.",
     "Reconciliation runs are long. Tokens expire inside them. Right now that means starting over.", "3"),

    ("Athan Z.", "Cofounder & CEO", "Copperlane", "copperlane.ai", "San Francisco", 5, "athan@copperlane.ai", "Verified",
     "AI reconciliation for high-volume payments",
     "High-volume + long-running + money. A partially-applied reconciliation is worse than none.",
     "A reconciliation run that dies halfway is worse than one that never started — you're now in an ambiguous state. Revive resumes exactly-once.", "3"),

    ("Wye H.", "CEO, Co-Founder", "Proximitty", "proximitty.ai", "San Francisco", 3, "wyeyew@proximitty.ai", "Verified",
     "AI operating system for commercial lending",
     "Lending workflows touch bank + doc + CRM systems and run for days.",
     "A multi-day lending workflow will outlive at least one of its access tokens. Today that's a restart; it doesn't have to be.", "3"),

    ("Zi Zhang", "Co-Founder & CTO", "Proximitty", "proximitty.ai", "United States", 3, "zi@proximitty.ai", "Verified",
     "AI operating system for commercial lending", "Second contact at Proximitty — technical.",
     "Technical framing: checkpoint + generation-based resume so a re-auth doesn't replay completed steps.", "3"),

    ("Clarence C.", "CEO / Co-Founder", "Travo", "travoai.com", "Stanford", 4, "clarence@travoai.com", "Verified",
     "AI-native internal audit automation",
     "Audit = long unattended evidence-collection runs across many systems.",
     "Audit runs collect evidence across a dozen systems over hours. One expired grant and the evidence set is quietly incomplete.", "3"),

    ("Aryah Oztanir", "Co-Founder / CEO", "o11", "o11.ai", "New York", 4, "aryah@o11.ai", "Verified",
     "PLG automation inside Microsoft 365 and Google Workspace",
     "Pure Workspace/M365 refresh-token dependency. The textbook case.",
     "You live entirely on M365 + Workspace refresh tokens. You already know what invalid_grant costs you. We turn it into a re-approval link.", "3"),

    # ---------- TIER 4: healthcare ops agents (EHR sessions die constantly; dropped run = dropped claim) ----------
    ("Mark P.", "CEO & Co-Founder", "Beacon Health", "beaconhealth.ai", "San Francisco", 5, "mark@beaconhealth.ai", "Verified",
     "AI healthcare operations agents for EHR workflows",
     "EHR sessions expire aggressively by design. A dropped run is a dropped claim, i.e. real money.",
     "EHR sessions expire by design. A dropped agent run is a dropped claim. That's the most expensive token death in software.", "4"),

    ("Meng S.", "CEO & Co-Founder", "Ruma Care", "rumacare.com", "San Francisco", 5, "ms@rumacare.com", "Verified",
     "Prior auth + reimbursement, EHR-integrated",
     "Prior-auth runs are long, high-value, and die on session expiry.",
     "A prior-auth run that dies at step 7 of 9 has to start over — and the payer window doesn't wait.", "4"),

    ("Akhil S.", "Founder & CEO", "Docura Health", "docurahealth.com", "San Francisco", 3, "akhil@docurahealth.com", "Verified",
     "AI medical-legal report writing for physicians",
     "EHR + doc-system grants held by physicians who will not troubleshoot auth.",
     "Physicians will not re-authorize anything. Revive routes the re-approval to whoever will and resumes the report where it stopped.", "4"),

    ("Emre Kaplaner", "Co-Founder", "Patientdesk.ai", "patientdesk.ai", "United States", 6, "emre@patientdesk.ai", "Verified",
     "AI receptionist for dental practices",
     "Practice-management OAuth held by a dental front desk. Silent death = missed patients.",
     "An AI receptionist that stops booking because a token expired = a phone nobody answers. Your customer finds out from an angry patient.", "4"),

    ("Adrian K.", "Co-Founder", "Mango Medical", "mangomedical.io", "San Francisco", 6, "adrian@mangomedical.io", "Verified",
     "AI medical billing with revenue-cycle automation",
     "Billing runs = money on the line, long-running, EHR/clearinghouse sessions.",
     "Revenue-cycle runs that die halfway leave claims in limbo. Resume-from-checkpoint is the difference between a retry and a write-off.", "4"),

    ("Nick W.", "Co-Founder, CEO", "MochaCare", "mochacare.com", "San Francisco", 2, "nick@mochacare.com", "Verified",
     "AI-powered home care operations",
     "Scheduling + EHR grants held by home-care coordinators.",
     "Home-care scheduling agents that go quiet don't error loudly — a caregiver just doesn't show up.", "4"),

    ("Sam Oberly", "CEO", "Scheduling Wizard", "schedulingwiz.com", "Philadelphia", 3, "sam@schedulingwiz.com", "Verified",
     "Medical residency schedule automation",
     "Long scheduling optimization runs against hospital systems.",
     "A scheduling run that dies mid-optimization is a wasted night. We resume from the checkpoint instead.", "4"),

    # ---------- TIER 5: vertical AI where a non-technical user holds the grant ----------
    ("Matthew Asir", "CEO", "LegalOS", "legalos.ai", "Fort Lauderdale", 8, "matthew@legalos.ai", "Verified",
     "AI-native immigration law / visa applications",
     "Government portals + doc systems. Session death mid-filing is catastrophic and deadline-bound.",
     "Immigration filings are deadline-bound. A run that dies at the portal login wall doesn't get a second chance at the deadline.", "5"),

    ("David E.", "Co-Founder, CEO", "Lexius", "lexius.ai", "San Francisco", 4, "david@lexius.ai", "Verified",
     "AI-native legal services", "Doc + email grants; long document-generation runs.",
     "Long doc-generation runs across client systems. One expired grant and the whole run restarts.", "5"),

    ("Oskar B.", "CEO", "Stilta", "stilta.com", "San Francisco", 7, "oskar@stilta.com", "Verified",
     "AI automation for HVAC job paperwork",
     "Field-service + accounting grants held by HVAC office staff.",
     "HVAC office staff will not reconnect an integration. They'll just go back to doing the paperwork by hand — and churn.", "5"),

    ("Jad B.", "Co-Founder", "Verdex", "verdexai.com", "San Francisco", 2, "jad@verdexai.com", "Verified",
     "AI-native review layer for real estate permits",
     "Municipal permit portals — the most hostile session environment there is.",
     "Municipal permit portals are the most hostile session environment in software. That's our home turf.", "5"),

    ("Evan R.", "Co-Founder", "Verdex", "verdexai.com", "San Francisco", 2, "evan@verdexai.com", "Verified",
     "AI-native review layer for real estate permits", "Second contact at Verdex.",
     "Technical framing of the permit-portal session-death pitch.", "5"),

    ("Omar E.", "Co-Founder & CTO", "Caretta", "caretta.so", "San Francisco", 4, "omar@caretta.so", "Verified",
     "Realtime AI for sales calls",
     "CRM + calendar + call grants; realtime means stale data is served confidently.",
     "Realtime sales AI on a stale CRM token serves confident, wrong context to a rep mid-call. Worse than silence.", "5"),
]

# Named founders with NO Apollo email — LinkedIn / site-contact route
NO_EMAIL = [
    ("Rithvik Vanga", "Co-founder, CEO", "Zatanna", "zatanna.ai", "Ann Arbor", 3,
     "AI agents for hardware procurement", "Apollo has no email on file. LinkedIn or site contact form."),
    ("George Z.", "CEO, Co-Founder", "AutoSitu", "autositu.com", "San Francisco", 3,
     "Agentic AI for fraud investigations", "Apollo has no email on file. LinkedIn."),
    ("Lyem N.", "Founder", "Asimov", "tryasimov.ai", "Berkeley", 5,
     "AI customer text layer for DTC brands", "Apollo has no email on file. LinkedIn."),
    ("Misha Baheti", "Founding Member", "Rubric AI", "therubric.ai", "New Delhi", 1,
     "AI agent optimization for production workflows", "Apollo has no email on file. LinkedIn."),
    ("Utkarsh Gill", "Co-Founder", "Traverse", "traverse.so", "Gurugram", 3,
     "AI workflow discovery", "Apollo has no email on file; co-founder Lance Y. is reachable at lance@traverse.so."),
    ("Ajay Misra", "President & Co-Founder", "o11", "o11.ai", "New York", 4,
     "PLG automation in M365 / Workspace", "Apollo has no email on file; CEO Aryah is reachable at aryah@o11.ai."),
    ("Moses Wayne", "Co-Founder", "Cofia", "cofia.ai", "New York", 2,
     "AI automations that observe workflows", "Co-founder Paola is reachable at paola@cofia.ai."),
    ("Tejas P.", "Co-Founder", "Glue", "buildwithglue.com", "San Francisco", 2,
     "AI brand operations", "Founder Perbhat is reachable at perbhat@buildwithglue.com."),
]

HDR = ["#", "Name", "Title", "Company", "Domain", "HQ", "Headcount", "Email", "Email status",
       "What they do", "Why they fit Revive", "Opening line (personalised)", "Tier", "Status"]

wb = Workbook()
ws = wb.active
ws.title = "Prospects"

HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY = Font(name="Arial", size=10)
FLAG = Font(name="Arial", size=10, color="C00000", bold=True)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TIER_FILL = {
    "1": PatternFill("solid", fgColor="E2EFDA"),
    "2": PatternFill("solid", fgColor="DDEBF7"),
    "3": PatternFill("solid", fgColor="FFF2CC"),
    "4": PatternFill("solid", fgColor="FCE4EC"),
    "5": PatternFill("solid", fgColor="EDEDED"),
}

ws.append(HDR)
for c in ws[1]:
    c.fill, c.font, c.border = HDR_FILL, HDR_FONT, BORDER
    c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

for i, r in enumerate(ROWS, start=1):
    name, title, comp, dom, hq, hc, email, est, what, why, hook, tier = r
    ws.append([i, name, title, comp, dom, hq, hc, email, est, what, why, hook, tier, "todo"])
    row = ws.max_row
    for c in ws[row]:
        c.font, c.border = BODY, BORDER
        c.alignment = Alignment(vertical="top", wrap_text=True)
    ws.cell(row=row, column=13).fill = TIER_FILL[tier]
    ws.cell(row=row, column=13).alignment = Alignment(horizontal="center", vertical="top")
    if "MISMATCH" in est:
        ws.cell(row=row, column=9).font = FLAG

start_no_email = ws.max_row + 2
ws.cell(row=start_no_email, column=1, value="NAMED FOUNDERS WITH NO APOLLO EMAIL — LinkedIn / site-contact route").font = Font(
    name="Arial", bold=True, size=10, color="1F3864")

ws.append(HDR)
for c in ws[ws.max_row]:
    c.fill, c.font, c.border = HDR_FILL, HDR_FONT, BORDER
    c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

for j, (name, title, comp, dom, hq, hc, what, note) in enumerate(NO_EMAIL, start=len(ROWS) + 1):
    ws.append([j, name, title, comp, dom, hq, hc, "—", "Not in Apollo", what, note, "", "—", "todo"])
    for c in ws[ws.max_row]:
        c.font, c.border = BODY, BORDER
        c.alignment = Alignment(vertical="top", wrap_text=True)

widths = [5, 20, 24, 18, 20, 16, 10, 26, 15, 40, 52, 66, 6, 10]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:N{len(ROWS)+1}"

# ---- Method / caveats sheet ----
m = wb.create_sheet("Method & caveats")
m.column_dimensions["A"].width = 26
m.column_dimensions["B"].width = 110
notes = [
    ("Built", "11 July 2026"),
    ("ICP filter", "Does a NON-TECHNICAL end user OAuth their own Gmail / QuickBooks / Shopify / EHR into this "
                   "product, and does the product silently die when that grant expires? Everything here passes that test. "
                   "'Uses AI' was not sufficient."),
    ("Persona", "Founder / CEO / CTO at <20 people. At this size the founder is both the buyer and the person "
                "debugging invalid_grant at 2am."),
    ("Company source", "YC W26 batch directory (via extruct.ai data room), cross-checked against Apollo headcount."),
    ("Contact source", "Apollo.io people-search + email reveal, run live in the browser on 11 July 2026. "
                       "75 lead credits consumed, 20 remaining."),
    ("Email status", "Every address in column H was returned by Apollo's reveal, not guessed. "
                     "No firstname@domain.com pattern-matching was used anywhere in this file."),
    ("", ""),
    ("⚠ CHECK BEFORE SENDING", "Sparkles (row 27): Apollo lists the CEO as 'Ai B------' but returned dan@sparkles.dev. "
                               "The name and the address don't agree. Verify on LinkedIn before you send, or you'll "
                               "open with the wrong name."),
    ("⚠ DROPPED", "'Klaus' (klausai.com) was dropped. Apollo matched it to warren@drinkklaus.com — a beverage company, "
                  "not the AI-employee startup. Wrong company; the email would have bounced or, worse, landed."),
    ("⚠ DROPPED", "Coffee (490 employees) and Reevo (110) were cut on your instruction — they fail the <20-person "
                  "filter and are big enough to build reauth in-house."),
    ("", ""),
    ("Not covered", "Fixture, Clice, Q2Q, Carly, Silas, BrowserAct, Noetic, IncidentFox, Canary had no "
                    "founder-level record in Apollo. Worth a manual LinkedIn pass."),
    ("LinkedIn URLs", "Not pulled — Apollo's Links column requires a separate scrape. Ask and I'll add them; it costs "
                      "no credits."),
    ("", ""),
    ("Suggested sequence", "1) Tier 1 first — they feel this daily and there are only ~2-6 people at each company, so "
                           "the founder reads their own email. 2) Tier 4 (healthcare) has the highest pain but the "
                           "slowest yes — great for design-partner feedback, bad for a fast logo. 3) Salus / Lucent / "
                           "Piris are partners, not customers: different pitch, don't sell them a seat."),
]
for k, v in notes:
    m.append([k, v])
for row in m.iter_rows():
    for c in row:
        c.font = Font(name="Arial", size=10, bold=(c.column == 1))
        c.alignment = Alignment(vertical="top", wrap_text=True)
        if isinstance(c.value, str) and c.value.startswith("⚠"):
            c.font = Font(name="Arial", size=10, bold=True, color="C00000")

wb.save("/sessions/happy-inspiring-darwin/mnt/Revive/Revive-alpha-customers.xlsx")
print(f"contacts with verified email: {len(ROWS)}")
print(f"named, no email: {len(NO_EMAIL)}")
print(f"total people: {len(ROWS) + len(NO_EMAIL)}")
print(f"unique companies: {len({r[3] for r in ROWS} | {n[2] for n in NO_EMAIL})}")
